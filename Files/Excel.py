from openpyxl.styles import PatternFill, Alignment
from FileManager import ExcelFile


class MeasurementTemplate:
    def __init__(self, file_path, sheet_name, data, mode='w'):
        self.file_path = file_path
        self.book = None
        self.sheet_name, self.sheet = sheet_name, None
        self.data = data
        self.file = None
        self.mode = mode

    def get_file(self):
        self.file = ExcelFile(self.file_path, self.mode)
        self.book = self.file.open()
        return self.book

    def create_sheet(self):
        self.sheet = self.file.open_sheet()
        return self.sheet

    def write_data(self):
        counter = 3
        i = 0
        for n in range(self.data['meas_num']):
            self.sheet.cell(row=2, column=counter).value = round(self.data['Current'][n][0])
            for h in range(15):
                self.sheet.cell(row=3 + h, column=counter).value = self.data['Ags'][n][h]
                self.sheet.cell(row=19 + h, column=counter).value = self.data['Bgs'][n][h]
                self.sheet.cell(row=35 + h, column=counter).value = self.data['Field'][n][h]
                self.sheet.cell(row=51 + h, column=counter).value = self.data['Relative_field'][n][h]
                self.sheet.cell(row=67 + h, column=counter).value = self.data['Relative_Ags'][n][h]
                self.sheet.cell(row=83 + h, column=counter).value = self.data['Relative_Bgs'][n][h]

            counter += 1
            i += 1

    def save_book(self):
        self.file.save()

    def style_sheet(self):
        self.sheet.column_dimensions['A'].width = 34
        self.sheet.cell(row=2, column=2).value = 'current [A]'

        self.sheet.merge_cells('A3:A17')
        self.sheet.cell(row=3, column=1).value = 'A Gs'
        self.sheet.cell(row=3, column=1).fill = PatternFill(start_color='0099CCFF', end_color='0099CCFF',
                                                                 fill_type='solid')
        self.sheet.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')

        self.sheet.merge_cells('A19:A33')
        self.sheet.cell(row=19, column=1).value = 'B Gs'
        self.sheet.cell(row=19, column=1).fill = PatternFill(start_color='00FF8080', end_color='00FF8080',
                                                                  fill_type='solid')
        self.sheet.cell(row=19, column=1).alignment = Alignment(horizontal='center', vertical='center')

        self.sheet.merge_cells('A35:A49')
        self.sheet.cell(row=35, column=1).value = 'Amplitude Gs'
        self.sheet.cell(row=35, column=1).alignment = Alignment(horizontal='center', vertical='center')

        self.sheet.merge_cells('A51:A65')
        self.sheet.cell(row=51, column=1).value = 'Relative'
        self.sheet.cell(row=51, column=1).alignment = Alignment(horizontal='center', vertical='center')

        self.sheet.merge_cells('A67:A81')
        self.sheet.cell(row=67, column=1).value = 'A relative Gs'
        self.sheet.cell(row=67, column=1).fill = PatternFill(start_color='0099CCFF', end_color='0099CCFF',
                                                                 fill_type='solid')
        self.sheet.cell(row=67, column=1).alignment = Alignment(horizontal='center', vertical='center')

        self.sheet.merge_cells('A83:A97')
        self.sheet.cell(row=83, column=1).value = 'B relative Gs'
        self.sheet.cell(row=83, column=1).fill = PatternFill(start_color='00FF8080', end_color='00FF8080',
                                                                  fill_type='solid')
        self.sheet.cell(row=83, column=1).alignment = Alignment(horizontal='center', vertical='center')

        for i in range(6):
            for h in range(15):
                self.sheet.cell(row=3 + h + i * 16, column=2).value = f'{h + 1}'

    def run(self):
        self.get_file()
        self.create_sheet()
        self.write_data()
        self.style_sheet()
        self.save_book()
